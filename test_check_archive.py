import os
from zipfile import ZipFile
import openpyxl
from PyPDF2 import PdfReader
import pandas as pd

def test_check_archived_txt():
    with ZipFile('tmp/testzip.zip') as zf:
        assert zf.read('Random_note.txt').decode('utf-8') == 'Just some random text'
        assert zf.getinfo('Random_note.txt').file_size == os.path.getsize('resources/Random_note.txt')

def test_check_archived_pdf():
    with ZipFile('tmp/testzip.zip') as zf:
        archived_pdf = PdfReader(zf.open('bar_menu.pdf'))
        non_archived_pdf = PdfReader('resources/bar_menu.pdf')
        assert len(archived_pdf.pages) == len(non_archived_pdf.pages)
        assert 'Кугель  с грибным соусом' in archived_pdf.pages[0].extract_text()
        assert zf.getinfo('bar_menu.pdf').file_size == os.path.getsize('resources/bar_menu.pdf')

def test_check_archived_xls():
    with ZipFile('tmp/testzip.zip') as zf:
        archived_book_read = pd.read_excel(zf.open('Company_staff.xls'))
        archived_book_excel = pd.ExcelFile(zf.open('Company_staff.xls'))
        non_archived_book_read = pd.read_excel('resources/Company_staff.xls')
        non_archived_book_excel = pd.ExcelFile('resources/Company_staff.xls')
        assert archived_book_read.columns.any() == non_archived_book_read.columns.any() and archived_book_read.columns.all() == non_archived_book_read.columns.all()
        assert archived_book_excel.sheet_names == non_archived_book_excel.sheet_names
        assert archived_book_read.shape[0] == non_archived_book_read.shape[0] and archived_book_read.shape[1] == non_archived_book_read.shape[1]
        assert zf.getinfo('Company_staff.xls').file_size == os.path.getsize('resources/Company_staff.xls')

def test_check_archived_xlsx():
    with ZipFile('tmp/testzip.zip') as zf:
        archived_book = openpyxl.load_workbook(zf.open('Office_characters.xlsx')).active
        non_archived_book = openpyxl.load_workbook('resources/Office_characters.xlsx').active
        assert archived_book.cell(row=2, column=2).value == non_archived_book.cell(row=2, column=2).value
        assert archived_book.max_column == non_archived_book.max_column
        assert archived_book.max_row == non_archived_book.max_row
        assert zf.getinfo('Office_characters.xlsx').file_size == os.path.getsize('resources/Office_characters.xlsx')

